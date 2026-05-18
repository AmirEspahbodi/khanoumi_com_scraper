from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path

logger = logging.getLogger("scraper")


# ---------------------------------------------------------------------------
# Search-query loader
# ---------------------------------------------------------------------------


def get_search_data(path: str = "Book1.xlsx") -> tuple[tuple[str, ...]]:
    """
    Read search queries from the *second column* of every row in *path*.
    Blank / whitespace-only cells are silently skipped.

    Raises FileNotFoundError if the workbook does not exist so the caller
    gets a clear error rather than a silent empty tuple.
    """
    xlsx_path = Path(path)
    if not xlsx_path.exists():
        raise FileNotFoundError(
            f"Search-query workbook not found: '{xlsx_path.resolve()}'. "
            "Create Book1.xlsx with queries in column B."
        )

    from openpyxl import load_workbook  # lazy import — only needed at startup

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    data: list[tuple[str]] = []
    counter = 0
    for row in ws.iter_rows(min_col=0, max_col=8, values_only=True):
        if counter == 0:
            counter += 1
            continue
        if row[5] is None or str(row[5]).strip().lower() in ["", "nan"]:
            continue
        data.append((row[1], row[2], row[3], row[4], row[5], row[6], row[7]))
    wb.close()

    if not data:
        logger.warning("get_search_data: no queries found in column B of '%s'.", path)

    logger.info("get_search_data: loaded %d queries from '%s'.", len(data), path)
    return tuple(data)


# ---------------------------------------------------------------------------
# Scraper configuration
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ScraperConfig:
    """Central configuration — tweak without touching business logic."""

    # Target
    base_url: str = "https://www.khanoumi.com/"
    search_input_xpath: str = '//*[@id="header"]/div[1]/div[1]/div[1]/form/div[1]/input'
    search_button_xpath: str = (
        '//*[@id="header"]/div[1]/div[1]/div[1]/form/div[1]/div[1]/svg'
    )

    # Xpath selectors (adapt to target site)
    sel_product_container_xpath: str = (
        '//div[@data-sentry-component="VirtualizedPLPGrid"]'
    )
    sel_next_page: str = '//a[@aria-label="Next page"]'
    check_no_product_container_xpath = (
        "//html/body/div[3]/div[1]/section[2]/div[3]/div[1]/img"
    )
    each_product_xpath = '//div[@data-sentry-component="VirtualizedPLPGrid"]/div/div'
    one_product_xpath = (
        '//div[@data-sentry-component="VirtualizedPLPGrid"]/div[1]/div[1]'
    )

    # Concurrency
    num_workers: int = 4
    queue_maxsize: int = 100000

    # Resilience
    max_retries: int = 3
    backoff_base: float = 2.0  # seconds (exponential: base^attempt)

    # Timing (seconds)
    min_delay: float = 0.8
    max_delay: float = 2.5
    navigation_timeout: int = 30_000  # ms
    element_timeout: int = 30_000  # ms

    # Stealth
    user_agents: tuple[str, ...] = (
        # Keep UA consistent with Chrome channel used below
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.6613.119 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.6533.120 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.6478.127 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.6422.142 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.6367.201 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.6261.112 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.6099.225 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.5993.118 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.5790.171 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.5615.138 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.5414.120 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.5359.125 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.5060.134 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.127 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.102 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.6613.120 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.6478.127 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.6367.201 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.6261.112 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.6099.234 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.5993.117 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.5845.187 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.5735.198 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.5481.177 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.5195.125 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.127 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.99 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.6613.119 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.6478.114 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.6367.201 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.6167.139 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.6045.159 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.5790.170 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.5615.165 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.5359.124 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.5005.115 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.6478.132 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.6261.137 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.5993.124 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.5735.239 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.5414.125 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.133 Safari/537.36",
    )
    # Bot-challenge fingerprints
    bot_challenge_keywords: tuple[str, ...] = (
        "cloudflare",
        "verify you are human",
        "ddos-guard",
        "just a moment",
        "access denied",
        "bot detection",
    )
    sel_title: str = "h1.product__title"
    sel_price: str = "span.product__price"
    sel_sku: str = "[data-sku], .product__sku"
    sel_availability: str = ".product__availability"


CFG = ScraperConfig()
