import os
import re
import time

from dotenv import load_dotenv
from google import genai
from google.genai import types
from openpyxl import load_workbook
from pydantic import BaseModel, ValidationError

# Load environment variables
load_dotenv(override=True)

# Configuration
INPUT_FILE = "Book1.xlsx"
COLUMN_NAME = "title"
MODEL_NAME = "gemini-3-flash-preview"


# Define the Pydantic schema for structured output
class NormalizedResult(BaseModel):
    normalized_search_query: str
    normalized_product_name1: str
    normalized_product_name2: str
    normalized_product_name3: str
    normalized_product_name4: str
    normalized_product_name5: str


class KeyManager:
    def __init__(self, keys):
        if not keys:
            raise ValueError("API Keys list is empty! Check your .env file format.")
        self.keys = keys
        self.current_index = 0
        self.client = genai.Client(api_key=self.keys[self.current_index])
        print(f"Initialized with {len(self.keys)} API keys. Starting with Key 1.")

    def get_client(self):
        return self.client

    def switch_key(self):
        self.current_index += 1
        # Reached the end of the key list
        if self.current_index >= len(self.keys):
            print(
                "\n[!] All API keys exhausted. Waiting for 1 minute before restarting cycle..."
            )
            time.sleep(60)
            self.current_index = 0

        print(f"\n[>] Switched to API Key {self.current_index + 1} of {len(self.keys)}")
        # Instantiate a new client with the new key
        self.client = genai.Client(api_key=self.keys[self.current_index])


def load_api_keys() -> list[str]:
    """
    Loads API keys from a comma-separated list in the environment variable.
    """
    # 1. Fetch the raw string from the environment
    raw_keys = os.getenv("GEMINI_API_KEYS", "")

    if not raw_keys:
        # Fallback to the singular old name just in case
        raw_keys = os.getenv("GEMINI_API_KEY", "")

    if not raw_keys:
        raise ValueError("GEMINI_API_KEYS environment variable is missing or empty.")

    # 2. Split, strip whitespace, and filter out empty strings
    keys = [key.strip() for key in raw_keys.split(",") if key.strip()]

    if not keys:
        raise ValueError("No valid API keys found after parsing.")

    return keys


# Initialize our KeyManager globally
try:
    API_KEYS = load_api_keys()
    key_manager = KeyManager(API_KEYS)
except ValueError as e:
    print(f"Configuration Error: {e}")
    exit(1)


def normalize_query(raw_query: str) -> NormalizedResult | None:
    prompt = f"""# System Role
You are an expert Data Normalizer and SEO Specialist in the cosmetics and beauty industry. Your exact task is to take messy, human-written cosmetic product names from an Excel file and convert them into one optimized search query and five specifically structured, normalized product names.

# Objective
For the provided raw product name, you must generate a JSON response exactly matching this schema:
{{
"normalized_search_query": "string",
"normalized_product_name1": "string",
"normalized_product_name2": "string",
"normalized_product_name3": "string",
"normalized_product_name4": "string",
"normalized_product_name5": "string"
}}

# Strict Rules for "normalized_search_query":
- Broad enough to yield results, specific enough to filter out junk.
- Include ONLY the Brand, Core Product Type, and Main Line/Model.
- EXCLUDE highly specific details like exact volume (e.g., 100ml) or specific color codes (e.g., NC41, m010) to maximize search hits.

# CRITICAL RULES FOR ALL "normalized_product_name" VARIATIONS (PREVENT FALSE POSITIVES):
- ZERO HALLUCINATION: DO NOT ADD any information. If the raw name lacks a brand, model, size, or color, DO NOT add one.
- NO DATA LOSS: DO NOT REMOVE any core information. Specific shades (e.g., NC41) or volumes (e.g., 125m) MUST be present in all 5 variations.
- Fix all obvious typos in all variations.
- Standardize volume/weight phrases to official Persian formats (e.g., "م" or "میل" becomes "میلی لیتر").

# Specific Rules for the 5 Variations:
- normalized_product_name1 (Standard Format): [Product Type] + [Brand (Persian)] + [Model/Feature] + [Code/Color] + [Volume (Standardized)]. Fix spacing.
- normalized_product_name2 (Mixed/Brand-First): [Brand (Persian/English)] + [Product Type] + [Model/Code (Keep original English characters if applicable)].
- normalized_product_name3 (Expanded/Cleaned): Reorder slightly to mimic formal catalog names. Keep all details.
- normalized_product_name4 (All-Persian Standard E-commerce): Base format MUST be Standard E-commerce. ALL brands, categories, and Finglish terms MUST be translated/transliterated to Persian script (e.g., MAC -> مک, Full Lash -> فول لش). Volume/Weight standardized in Persian format.
- normalized_product_name5 (English-Brand Standard E-commerce): Base text structure MUST be Persian (e.g., کرم پودر ... مدل ... حجم ...). ONLY Brands, Marks, and Finglish/Specific terms MUST be translated to English (e.g., مک -> MAC, فول لش -> Full Lash). Volume/Weight standardized in Persian format.

# Output Format
You must respond ONLY with a valid JSON object. Do not include markdown tags like ```json or any conversational text.

# Examples
Input: "مک کرم پودر استودیو فیکسNC41"
Output:
{{
"normalized_search_query": "کرم پودر مک استودیو فیکس",
"normalized_product_name1": "کرم پودر مک مدل استودیو فیکس رنگ NC41",
"normalized_product_name2": "مک کرم پودر Studio Fix شماره NC41",
"normalized_product_name3": "کرم پودر مک استودیو فیکس NC41",
"normalized_product_name4": "کرم پودر مک مدل استودیو فیکس رنگ ان سی 41",
"normalized_product_name5": "کرم پودر MAC مدل Studio Fix رنگ NC41"
}}

Input: "لورال پرایمر تیوپی اینفالیبل 35 میل"
Output:
{{
"normalized_search_query": "پرایمر تیوپی لورال اینفالیبل",
"normalized_product_name1": "پرایمر تیوپی لورال مدل اینفالیبل حجم 35 میلی لیتر",
"normalized_product_name2": "لورال پرایمر Infallible حجم 35 میل",
"normalized_product_name3": "پرایمر تیوپی لورال اینفالیبل 35 میلی لیتر",
"normalized_product_name4": "پرایمر تیوپی لورآل مدل اینفالیبل حجم 35 میلی لیتر",
"normalized_product_name5": "پرایمر تیوپی L'Oreal مدل Infallible حجم 35 میلی لیتر"
}}

Input: "بل ریمل فول لش"
Output:
{{
"normalized_search_query": "ریمل بل فول لش",
"normalized_product_name1": "ریمل بل مدل فول لش",
"normalized_product_name2": "بل ریمل Full Lash",
"normalized_product_name3": "ریمل چشم بل فول لش",
"normalized_product_name4": "ریمل بل مدل فول لش",
"normalized_product_name5": "ریمل Bell مدل Full Lash"
}}

Input: "ادکلن دیویدوف کول واتر 125م"
Output:
{{
"normalized_search_query": "ادکلن دیویدوف کول واتر",
"normalized_product_name1": "ادکلن دیویدوف مدل کول واتر حجم 125 میلی لیتر",
"normalized_product_name2": "دیویدوف ادکلن Cool Water حجم 125 میل",
"normalized_product_name3": "ادکلن دیویدوف کول واتر 125 میلی لیتر",
"normalized_product_name4": "ادکلن دیویدوف مدل کول واتر حجم 125 میلی لیتر",
"normalized_product_name5": "ادکلن Davidoff مدل Cool Water حجم 125 میلی لیتر"
}}

# Real Task
Process the following raw product name exactly according to the rules and return the JSON:
Input: "{raw_query}"
"""

    while True:
        try:
            client = key_manager.get_client()

            # Use structured output config to enforce the JSON schema
            response = client.models.generate_content(
                model=MODEL_NAME,
                contents=prompt,
                config=types.GenerateContentConfig(
                    response_mime_type="application/json",
                    response_schema=NormalizedResult,
                    temperature=0.1,
                ),
            )

            # Parse and validate the response automatically using Pydantic
            parsed_result = NormalizedResult.model_validate_json(response.text)
            return parsed_result

        except ValidationError as ve:
            print(f"Pydantic Validation error for '{raw_query}': {ve}")
            return None
        except Exception as e:
            error_msg = str(e).lower()

            # Catch 429 Too Many Requests, 403 Quota Exceeded, or resource exhaustion
            if any(
                keyword in error_msg
                for keyword in [
                    "429",
                    "403",
                    "quota",
                    "exhausted",
                    "rate limit",
                    "too many requests",
                ]
            ):
                print(f"Rate limit or quota hit (Error: {e}). Switching API key...")
                key_manager.switch_key()
                continue
            else:
                # If it's a different kind of error, log it and move to the next item
                print(f"Error processing '{raw_query}': {e}")
                return None


def main():
    if not os.path.exists(INPUT_FILE):
        print(f"Error: {INPUT_FILE} not found.")
        return

    # Load the workbook and get the active worksheet
    wb = load_workbook(INPUT_FILE)
    ws = wb.active

    # Extract headers from the first row to find column indices
    headers = [cell.value for cell in ws[1]]

    if COLUMN_NAME not in headers:
        print(f"Error: Column '{COLUMN_NAME}' not found in the Excel file.")
        return

    # openpyxl uses 1-based indexing for columns
    title_col_idx = headers.index(COLUMN_NAME) + 1

    # Define the required output columns
    output_columns = [
        "normalized_search_query",
        "normalized_product_name1",
        "normalized_product_name2",
        "normalized_product_name3",
        "normalized_product_name4",
        "normalized_product_name5",
    ]

    col_indices = {}

    # Ensure all 4 columns exist in the header, otherwise append them
    for col_name in output_columns:
        if col_name in headers:
            col_indices[col_name] = headers.index(col_name) + 1
        else:
            new_col_idx = len(headers) + 1
            ws.cell(row=1, column=new_col_idx, value=col_name)
            headers.append(col_name)  # Keep track dynamically
            col_indices[col_name] = new_col_idx

    wb.save(INPUT_FILE)  # Save headers immediately

    total_rows = ws.max_row - 1  # Subtract 1 for the header row
    print(f"Start processing {total_rows} rows...\n")

    # Iterate starting from row 2 (skipping header)
    for row_num in range(2, ws.max_row + 1):
        # Check if the row has already been processed successfully
        already_processed = True
        for col_name in output_columns:
            cell_val = ws.cell(row=row_num, column=col_indices[col_name]).value
            if cell_val is None or str(cell_val).strip().lower() in ["", "nan"]:
                already_processed = False
                break

        if already_processed:
            continue

        title_cell = ws.cell(row=row_num, column=title_col_idx)
        raw_q = title_cell.value

        # Skip if the source title is empty
        if raw_q is None or str(raw_q).strip() == "":
            continue

        raw_q_str = str(raw_q)
        print(f"Processing ({row_num - 1}/{total_rows}): {raw_q_str}")

        # Fetch using the Pydantic/Gemini integration
        normalized_result = normalize_query(raw_q_str)

        if normalized_result:
            # Map Pydantic model values directly back into the specific Excel cells
            ws.cell(
                row=row_num,
                column=col_indices["normalized_search_query"],
                value=normalized_result.normalized_search_query,
            )
            ws.cell(
                row=row_num,
                column=col_indices["normalized_product_name1"],
                value=normalized_result.normalized_product_name1,
            )
            ws.cell(
                row=row_num,
                column=col_indices["normalized_product_name2"],
                value=normalized_result.normalized_product_name2,
            )
            ws.cell(
                row=row_num,
                column=col_indices["normalized_product_name3"],
                value=normalized_result.normalized_product_name3,
            )
            ws.cell(
                row=row_num,
                column=col_indices["normalized_product_name4"],
                value=normalized_result.normalized_product_name4,
            )
            ws.cell(
                row=row_num,
                column=col_indices["normalized_product_name5"],
                value=normalized_result.normalized_product_name5,
            )
            # Overwrite the input file to save progress safely
            wb.save(INPUT_FILE)
            print("Saved outputs successfully.")

        # Sleep to respect base rate limits
        time.sleep(2)


if __name__ == "__main__":
    main()
